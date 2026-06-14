"""
HỆ THỐNG THU THẬP BÁO CÁO TGDV - PHIÊN BẢN V5.0
Tối ưu: Caching toàn diện, UI hành chính trang nhã, st.secrets cho thông tin nhạy cảm
"""

import streamlit as st
import pandas as pd
import json
import os
import hashlib
import plotly.express as px
import plotly.graph_objects as go
import io
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================
# CẤU HÌNH TRANG
# ==========================================
st.set_page_config(
    page_title="Hệ thống Báo cáo TGDV",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ------------------------------------------
# GIAO DIỆN: Bảng màu hành chính trang nhã
# Nền: xanh navy đậm (#1F3A5F) - vàng đồng nhấn (#B08D57) - xám trắng nền (#F5F6F8)
# ------------------------------------------
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Be+Vietnam+Pro:wght@300;400;500;600;700;800;900&display=swap');

    html, body, [class*="css"] { font-family: 'Be Vietnam Pro', sans-serif; }
    .stApp { background-color: #F5F6F8; }
    .block-container { padding-top: 1.5rem; padding-bottom: 2rem; }

    .main-header {
        background: linear-gradient(135deg, #1F3A5F 0%, #2C4A75 100%);
        color: #F5F6F8; font-weight: 800; text-align: center; text-transform: uppercase;
        letter-spacing: 1.5px; padding: 20px 30px; border-radius: 10px; margin-bottom: 26px;
        font-size: 1.25rem; box-shadow: 0 4px 18px rgba(31,58,95,0.18);
        border-bottom: 3px solid #B08D57;
    }
    .main-header span { color: #D9B26A; }

    .metric-row { display: flex; gap: 16px; margin-bottom: 24px; }
    .metric-card {
        flex: 1; background: white; padding: 18px 16px; border-radius: 10px;
        border-left: 4px solid #1F3A5F; box-shadow: 0 1px 8px rgba(0,0,0,0.05); text-align: center;
    }
    .metric-card.red   { border-left-color: #A0443A; }
    .metric-card.green { border-left-color: #3E7D5A; }
    .metric-card.gold  { border-left-color: #B08D57; }
    .metric-title  { font-size: 11px; color: #7A8694; font-weight: 700; text-transform: uppercase; letter-spacing: 1px; margin-bottom: 6px; }
    .metric-number { font-size: 30px; color: #1F3A5F; font-weight: 800; margin: 0; line-height: 1.1; }
    .metric-number.red   { color: #A0443A; }
    .metric-number.green { color: #3E7D5A; }
    .metric-number.gold  { color: #B08D57; }
    .metric-sub { font-size: 11px; color: #A4ADB8; margin-top: 3px; }

    .section-title {
        display: flex; align-items: center; gap: 10px; color: #1F3A5F; font-weight: 700; font-size: 0.98rem;
        text-transform: uppercase; letter-spacing: 0.4px; border-left: 4px solid #B08D57; padding-left: 12px;
        margin: 26px 0 14px 0;
    }

    [data-testid="stForm"] { background-color: #ffffff; padding: 22px; border-radius: 10px; border: 1px solid #E3E7EC; box-shadow: 0 1px 10px rgba(0,0,0,0.04); }
    [data-testid="stExpander"] { background-color: #FAFBFC !important; border: 1px solid #E3E7EC !important; border-radius: 8px !important; margin-bottom: 10px !important; }
    [data-testid="stExpander"] summary { font-weight: 700 !important; color: #1F3A5F !important; }

    .stButton>button { background-color: #1F3A5F; color: white; font-weight: 600; border-radius: 6px; border: none; padding: 9px 18px; transition: all 0.2s; }
    .stButton>button:hover { background-color: #B08D57; color: white; }

    /* ---- SIDEBAR ---- */
    [data-testid="stSidebar"] { background: linear-gradient(180deg, #1A3253 0%, #1F3A5F 100%); }
    [data-testid="stSidebar"] * { color: white !important; }
    [data-testid="stSidebar"] div[data-baseweb="select"] * { color: #1F3A5F !important; font-weight: 600; }
    [data-testid="stSidebar"] .stSelectbox label { color: #A9BCD4 !important; font-weight: 600; font-size: 12px; }
    [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] { color: #A9BCD4 !important; }
    [data-testid="stSidebar"] .stButton>button { background: rgba(255,255,255,0.12); border: 1px solid rgba(255,255,255,0.25); }
    [data-testid="stSidebar"] .stButton>button:hover { background: #B08D57; border-color: #B08D57; }

    /* VÁ LỖI HIỂN THỊ Ô NHẬP LIỆU Ở FORM ĐỔI MẬT KHẨU SIDEBAR */
    [data-testid="stSidebar"] [data-testid="stExpander"] label,
    [data-testid="stSidebar"] [data-testid="stExpander"] p { color: #1F3A5F !important; font-weight: 700; margin-bottom: 4px; }

    [data-testid="stSidebar"] div[data-baseweb="input"] {
        background-color: #FFFFFF !important;
        border: 1.5px solid #1F3A5F !important;
        border-radius: 6px !important;
    }

    [data-testid="stSidebar"] input {
        color: #1F3A5F !important;
        font-weight: 600 !important;
        -webkit-text-fill-color: #1F3A5F !important;
        background-color: transparent !important;
    }

    .badge { display: inline-block; padding: 3px 10px; border-radius: 99px; font-size: 11px; font-weight: 700; letter-spacing: 0.5px; }
    .badge-success { background: #E1EFE6; color: #3E7D5A; }
    .badge-danger  { background: #F5E4E1; color: #A0443A; }

    .progress-wrap { background: #E3E7EC; border-radius: 99px; height: 10px; overflow: hidden; margin: 8px 0; }
    .progress-fill { background: linear-gradient(90deg, #1F3A5F, #3E7D5A); height: 100%; border-radius: 99px; transition: width 0.6s ease; }

    .info-box { background: #EEF2F7; border: 1px solid #D6E0EC; border-left: 4px solid #1F3A5F; border-radius: 6px; padding: 12px 16px; font-size: 13px; color: #1F3A5F; margin-bottom: 16px; }
    .warning-box { background: #FBF4E9; border: 1px solid #EBDCC2; border-left: 4px solid #B08D57; border-radius: 6px; padding: 12px 16px; font-size: 13px; color: #8A6D3B; margin-bottom: 16px; }
    .success-box { background: #EAF5EE; border: 1px solid #CDE6D6; border-left: 4px solid #3E7D5A; border-radius: 6px; padding: 12px 16px; font-size: 13px; color: #2F6347; margin-bottom: 16px; }

    .stTabs [data-baseweb="tab-list"] { gap: 6px; background: white; padding: 6px; border-radius: 8px; box-shadow: 0 1px 6px rgba(0,0,0,0.04); }
    .stTabs [data-baseweb="tab"] { border-radius: 6px !important; font-weight: 600 !important; padding: 8px 18px !important; color: #7A8694 !important; }
    .stTabs [aria-selected="true"] { background: #1F3A5F !important; color: white !important; }

    .stNumberInput label, .stTextInput label, .stSelectbox label, .stTextArea label { font-weight: 600 !important; font-size: 13px !important; color: #3C4856 !important; }
    .stDataFrame { border-radius: 8px; overflow: hidden; }
    hr { border: none; border-top: 1px solid #E3E7EC; margin: 20px 0; }

    .login-wrap { max-width: 420px; margin: 60px auto; background: white; border-radius: 14px; padding: 40px 36px; box-shadow: 0 8px 30px rgba(0,0,0,0.08); border-top: 4px solid #1F3A5F; }
    .login-logo { text-align: center; margin-bottom: 28px; }
    .login-logo h2 { color: #1F3A5F; font-weight: 800; font-size: 1.25rem; margin: 8px 0 4px; }
    .login-logo p  { color: #7A8694; font-size: 13px; }

    input:disabled { background-color: #F5F6F8 !important; color: #1F3A5F !important; font-weight: 700; opacity: 1; -webkit-text-fill-color: #1F3A5F !important; }
</style>
""", unsafe_allow_html=True)

# ==========================================
# CẤU HÌNH DỮ LIỆU
# ==========================================
DATA_FILE   = "dulieu_baocao.json"
CONFIG_FILE = "config_donvi.json"
PASS_FILE   = "passwords.json"

DEFAULT_UNITS = [
    "Đảng ủy Công an tỉnh", "Đảng ủy Quân sự tỉnh", "Đảng ủy các cơ quan Đảng tỉnh", "Đảng ủy Ủy ban nhân dân tỉnh",
    "Phường Mỹ Lâm", "Phường Minh Xuân", "Phường Nông Tiến", "Phường An Tường", "Phường Bình Thuận", "Phường Hà Giang 1", "Phường Hà Giang 2",
    "Xã Thượng Lâm", "Xã Lâm Bình", "Xã Minh Quang", "Xã Bình An", "Xã Côn Lôn", "Xã Yên Hoa", "Xã Thượng Nông", "Xã Hồng Thái", "Xã Nà Hang", "Xã Tân Mỹ", "Xã Yên Lập", "Xã Tân An", "Xã Chiêm Hóa", "Xã Hòa An", "Xã Kiên Đài", "Xã Tri Phú", "Xã Kim Bình", "Xã Yên Nguyên", "Xã Yên Phú", "Xã Bạch Xa", "Xã Phù Lưu", "Xã Hàm Yên", "Xã Bình Xa", "Xã Thái Sơn", "Xã Thái Hòa", "Xã Hùng Lợi", "Xã Trung Sơn", "Xã Thái Bình", "Xã Tân Long", "Xã Xuân Vân", "Xã Lực Hành", "Xã Yên Sơn", "Xã Nhữ Khê", "Xã Tân Trào", "Xã Minh Thanh", "Xã Sơn Dương", "Xã Bình Ca", "Xã Tân Thanh", "Xã Sơn Thủy", "Xã Phú Lương", "Xã Trường Sinh", "Xã Hồng Sơn", "Xã Đông Thọ",
    "Xã Lũng Cú", "Xã Đồng Văn", "Xã Sà Phìn", "Xã Phố Bảng", "Xã Lũng Phìn", "Xã Sủng Máng", "Xã Sơn Vĩ", "Xã Mèo Vạc", "Xã Khâu Vai", "Xã Niêm Sơn", "Xã Tát Ngà", "Xã Thắng Mố", "Xã Bạch Đích", "Xã Yên Minh", "Xã Mậu Duệ", "Xã Du Già", "Xã Đường Thượng", "Xã Lùng Tám", "Xã Cán Tỷ", "Xã Nghĩa Thuận", "Xã Quản Bạ", "Xã Tùng Vài", "Xã Yên Cường", "Xã Đường Hồng", "Xã Bắc Mê", "Xã Minh Ngọc", "Xã Ngọc Đường", "Xã Lao Chải", "Xã Thanh Thủy", "Xã Phú Linh", "Xã Linh Hồ", "Xã Bạch Ngọc", "Xã Vị Xuyên", "Xã Việt Lâm", "Xã Tân Quang", "Xã Đồng Tâm", "Xã Liên Hiệp", "Xã Bằng Hành", "Xã Bắc Quang", "Xã Hùng An", "Xã Vĩnh Tuy", "Xã Đồng Yên", "Xã Tiên Yên", "Xã Xuân Giang", "Xã Bằng Lang", "Xã Yên Thành", "Xã Quang Bình", "Xã Tân Trịnh", "Xã Thông Nguyên", "Xã Hồ Thầu", "Xã Nậm Dịch", "Xã Tân Tiến", "Xã Hoàng Su Phì", "Xã Thàng Tín", "Xã Bản Máy", "Xã Pờ Ly Ngài", "Xã Xín Mần", "Xã Pà Vầy Sủ", "Xã Nấm Dẩn", "Xã Trung Thịnh", "Xã Khuôn Lùng", "Xã Trung Hà", "Xã Kiến Thiết", "Xã Hùng Đức", "Xã Minh Sơn", "Xã Minh Tân", "Xã Thuận Hòa", "Xã Tùng Bá", "Xã Thượng Sơn", "Xã Cao Bồ", "Xã Ngọc Long", "Xã Giáp Trung", "Xã Tiên Nguyên", "Xã Quảng Nguyên"
]

DANH_SACH_THANG = [f"Tháng {i}" for i in range(1, 13)]

SCHEMA = [
    ("don_vi",               "Đơn vị báo cáo",              "info",   "text"),
    ("nguoi_bao_cao",        "Người BC / SĐT",               "info",   "text"),
    ("ky_bao_cao",           "Kỳ báo cáo",                   "info",   "text"),
    ("ngay_nop",             "Thời điểm nộp",                "info",   "text"),
    ("ld_vanban",            "VB cấp ủy ban hành",           "ld",     "int"),
    ("ld_thammuu",           "VB tham mưu cấp trên",         "ld",     "int"),
    ("ld_cuochop",           "Cuộc họp, hội nghị",           "ld",     "int"),
    ("nq_hoinghi",           "Số hội nghị NQ",               "nq",     "int"),
    ("nq_nguoi",             "Số người tham gia NQ",         "nq",     "int"),
    ("nq_vanban",            "Số VB đã triển khai",          "nq",     "int"),
    ("nq_tyle",              "Tỷ lệ ĐV tham gia (%)",        "nq",     "float"),
    ("tt_tinbai",            "Số tin, bài, pano",            "tt",     "int"),
    ("tt_loa",               "Lượt loa truyền thanh",        "tt",     "int"),
    ("tt_buoi",              "Số buổi TT miệng",             "tt",     "int"),
    ("tt_nguoi",             "Số người nghe TT",             "tt",     "int"),
    ("tt_mxh_bai",           "Bài trên MXH/Cổng TT",         "tt",     "int"),
    ("tt_mxh_tuongtac",      "Lượt tương tác MXH",           "tt",     "int"),
    ("dl_baocao",            "Số BC DLXH gửi đi",            "dl",     "int"),
    ("dl_vande",             "Số vấn đề nổi cộm",            "dl",     "int"),
    ("dl_xuly",              "Số vụ việc đã xử lý",          "dl",     "int"),
    ("kg_chuongtrinh",       "Số CT tuyên truyền GD",        "kg",     "int"),
    ("kg_lop",               "Số buổi Y tế/Môi trường",      "kg",     "int"),
    ("kg_bd_chuyennghiep",   "Biểu diễn NT chuyên nghiệp",  "kg",     "int"),
    ("kg_bd_quanchung",      "Biểu diễn NT quần chúng",     "kg",     "int"),
    ("kg_clb_thanhlap",      "Số CLB VH-NT thành lập",      "kg",     "int"),
    ("kg_clb_thanhvien",     "Số thành viên CLB",           "kg",     "int"),
    ("kg_hd_vhtt",           "Số HĐ Lễ hội, Thể thao",      "kg",     "int"),
    ("kg_hoatdong",          "Số HĐ Văn hóa-Văn nghệ",      "kg",     "int"),
    ("kg_khokhan",           "Khó khăn KG, VH-VN",          "kg",     "text"),
    ("dv_mh_dangky",         "Mô hình DVK đăng ký",         "dv",     "int"),
    ("dv_mh_hieuqua",        "Mô hình DVK hiệu quả",        "dv",     "int"),
    ("dv_mh_moi",            "Mô hình mới trong kỳ",        "dv",     "int"),
    ("dv_cuocvandong",       "Số cuộc vận động, TT",        "dv",     "int"),
    ("dv_nguoithamgia",      "Số lượt người tham gia",      "dv",     "int"),
    ("dv_tiepxuc",           "Số buổi đối thoại ND",        "dv",     "int"),
    ("nv_duocgiao",          "Nhiệm vụ TT được giao",       "nv",     "int"),
    ("nv_hoanthanh",         "Nhiệm vụ TT hoàn thành",      "nv",     "int"),
    ("nv_dangtrienkhai",     "Nhiệm vụ đang triển khai",    "nv",     "int"),
    ("nv_ketqua",            "Kết quả nổi bật",             "nv",     "text"),
    ("bd_tinbai",            "Số tin bài CĐS",              "bd",     "int"),
    ("bd_cuocthi",           "Số cuộc thi CĐS",             "bd",     "int"),
    ("kq_tocongnghe",        "Số Tổ công nghệ số",          "bd",     "int"),
    ("ts_chibo",             "Tổng số chi bộ",              "bd",     "int"),
    ("kq_chibo_cd",          "Chi bộ SH chuyên đề số",      "bd",     "int"),
    ("kq_chibo_sotay",       "Chi bộ dùng Sổ tay ĐV",      "bd",     "int"),
    ("ts_cbccvc",            "Tổng số CBCCVC",              "bd",     "int"),
    ("kq_cb_ai",             "CB biết dùng AI",             "bd",     "int"),
    ("kq_cb_khoahoc",        "CB hoàn thành CĐS",           "bd",     "int"),
    ("ts_nd_truongthanh",    "Tổng ND trưởng thành",        "bd",     "int"),
    ("kq_nd_kynang",         "ND có Kỹ năng số",            "bd",     "int"),
    ("kq_nd_vneid",          "ND phổ cập VNeID",            "bd",     "int"),
    ("kq_nd_smartphone",     "ND dùng Smartphone",          "bd",     "int"),
    ("kq_lop_nd",            "Buổi học cộng đồng",          "bd",     "int"),
    ("tl_mohinh",            "Mô hình hay, sáng tạo",       "tl",     "text"),
    ("tl_khokhan",           "Khó khăn, vướng mắc chung",   "tl",     "text"),
]

GROUP_LABELS = {
    "info": ("THÔNG TIN CHUNG",              "1F3A5F"),
    "ld":   ("1. LÃNH ĐẠO, CHỈ ĐẠO",         "A0443A"),
    "nq":   ("2. QUÁN TRIỆT NGHỊ QUYẾT",     "2C4A75"),
    "tt":   ("3. CÔNG TÁC TUYÊN TRUYỀN",     "A0443A"),
    "dl":   ("4. DƯ LUẬN XÃ HỘI",            "2C4A75"),
    "kg":   ("5. KHOA GIÁO, VH-VN",          "A0443A"),
    "dv":   ("6. DÂN VẬN KHÉO",              "2C4A75"),
    "nv":   ("7. NHIỆM VỤ TRỌNG TÂM",        "A0443A"),
    "bd":   ("8. BÌNH DÂN HỌC VỤ SỐ",       "2C4A75"),
    "tl":   ("9. ĐÁNH GIÁ CHUNG",            "A0443A"),
}

ALL_KEYS  = [s[0] for s in SCHEMA]
KEY_LABEL = {s[0]: s[1] for s in SCHEMA}
KEY_GROUP = {s[0]: s[2] for s in SCHEMA}
KEY_TYPE  = {s[0]: s[3] for s in SCHEMA}
NUM_KEYS  = [s[0] for s in SCHEMA if s[3] in ("int", "float")]
TEXT_KEYS = [s[0] for s in SCHEMA if s[3] == "text"]


# ==========================================
# HÀM TIỆN ÍCH QUẢN LÝ DỮ LIỆU (CÓ CACHING)
# ==========================================

def _file_sig(path):
    """Trả về chữ ký (mtime, size) của file để dùng làm cache key — tránh đọc lại
    file khi nội dung chưa đổi nhưng vẫn tự refresh khi file thay đổi."""
    try:
        st_ = os.stat(path)
        return (st_.st_mtime_ns, st_.st_size)
    except FileNotFoundError:
        return (0, 0)


@st.cache_data(show_spinner=False)
def _read_json_cached(path, _sig):
    """Đọc JSON, cache theo chữ ký file. _sig chỉ dùng để invalidate cache."""
    try:
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return None


def load_data():
    data = _read_json_cached(DATA_FILE, _file_sig(DATA_FILE))
    return data if data is not None else []


def save_data(data):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)
    _read_json_cached.clear()


def load_units():
    u_list = _read_json_cached(CONFIG_FILE, _file_sig(CONFIG_FILE))
    if u_list is not None and len(u_list) >= 120:
        return u_list
    save_units(DEFAULT_UNITS)
    return DEFAULT_UNITS


def save_units(units):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(units, f, ensure_ascii=False, indent=4)
    _read_json_cached.clear()


def load_passwords():
    units = load_units()
    defaults = {u: f"TGDV@{i+1:03d}" for i, u in enumerate(units)}
    saved = _read_json_cached(PASS_FILE, _file_sig(PASS_FILE))
    if saved is not None:
        changed = False
        for u in units:
            if u not in saved:
                saved[u] = defaults[u]
                changed = True
        return saved
    return defaults


def save_passwords(pwds):
    with open(PASS_FILE, "w", encoding="utf-8") as f:
        json.dump(pwds, f, ensure_ascii=False, indent=4)
    _read_json_cached.clear()


def get_months_for_filter(filter_type):
    mapping = {
        "Quý I":              ["Tháng 1", "Tháng 2", "Tháng 3"],
        "Quý II":             ["Tháng 4", "Tháng 5", "Tháng 6"],
        "Quý III":            ["Tháng 7", "Tháng 8", "Tháng 9"],
        "Quý IV":             ["Tháng 10", "Tháng 11", "Tháng 12"],
        "6 Tháng Đầu Năm":   [f"Tháng {i}" for i in range(1, 7)],
        "6 Tháng Cuối Năm":  [f"Tháng {i}" for i in range(7, 13)],
        "9 Tháng":            [f"Tháng {i}" for i in range(1, 10)],
        "Cả Năm":             [f"Tháng {i}" for i in range(1, 13)],
    }
    return mapping.get(filter_type, DANH_SACH_THANG)


def get_old_val(old, key, default=None):
    val = old.get(key, default)
    if val is None:
        return 0 if KEY_TYPE.get(key) in ("int", "float") else ""
    if KEY_TYPE.get(key) == "int":
        try: return int(val)
        except: return 0
    if KEY_TYPE.get(key) == "float":
        try: return float(val)
        except: return 0.0
    return str(val) if val else ""


def text_agg(series):
    items = list(dict.fromkeys([str(v).strip() for v in series if pd.notna(v) and str(v).strip()]))
    return "\n".join([f"• {t}" for t in items]) if items else ""


@st.cache_data(show_spinner=False)
def build_summary_df_cached(records_json, ky_label):
    """records_json: JSON string của list bản ghi (đã lọc theo kỳ) — dùng làm cache key."""
    df_cur = pd.DataFrame(json.loads(records_json))
    return build_summary_df(df_cur, ky_label)


def build_summary_df(df_cur, ky_label):
    if df_cur.empty:
        return pd.DataFrame()
    for key in ALL_KEYS:
        if key not in df_cur.columns:
            df_cur[key] = 0 if KEY_TYPE.get(key) in ("int", "float") else ""

    agg_dict = {}
    for key in NUM_KEYS:
        if key in df_cur.columns:
            agg_dict[key] = "mean" if key == "nq_tyle" else "sum"
    for key in TEXT_KEYS:
        if key in df_cur.columns and key not in ("don_vi", "ky_bao_cao"):
            agg_dict[key] = text_agg
    agg_dict["nguoi_bao_cao"] = lambda x: ", ".join(dict.fromkeys([str(v).strip() for v in x if pd.notna(v)]))
    agg_dict["ngay_nop"]      = lambda x: ", ".join(dict.fromkeys([str(v).strip() for v in x if pd.notna(v)]))

    df_sum = df_cur.groupby("don_vi", as_index=False).agg(agg_dict)
    df_sum["ky_bao_cao"] = ky_label

    total = {}
    for col in df_sum.columns:
        if col == "don_vi":
            total[col] = "TỔNG CỘNG"
        elif col in NUM_KEYS and pd.api.types.is_numeric_dtype(df_sum[col]):
            total[col] = round(df_sum[col].mean(), 1) if col == "nq_tyle" else df_sum[col].sum()
        else:
            total[col] = ""
    df_sum = pd.concat([df_sum, pd.DataFrame([total])], ignore_index=True)
    return df_sum


@st.cache_data(show_spinner=False)
def build_excel_cached(df_sum_json, ky_label):
    df_sum = pd.read_json(io.StringIO(df_sum_json), orient="split")
    buf = build_excel(df_sum, ky_label)
    return buf.getvalue()


def build_excel(df_sum, ky_label):
    out_keys = [k for k in ALL_KEYS if k != "ky_bao_cao"]
    out_keys.insert(1, "ky_bao_cao")
    display_cols = [KEY_LABEL.get(k, k) for k in out_keys]

    df_ex = pd.DataFrame()
    for k, label in zip(out_keys, display_cols):
        df_ex[label] = df_sum[k] if k in df_sum.columns else ""

    super_headers = []
    cur_group = None
    start_col = 1
    for i, k in enumerate(out_keys):
        grp = "info" if k in ("don_vi", "ky_bao_cao", "nguoi_bao_cao", "ngay_nop") else KEY_GROUP.get(k, "tl")
        if cur_group is None:
            cur_group = grp; start_col = i + 1
        elif grp != cur_group:
            lbl, color = GROUP_LABELS.get(cur_group, ("KHÁC", "666666"))
            super_headers.append((start_col, i, lbl, color))
            cur_group = grp; start_col = i + 1
    if cur_group:
        lbl, color = GROUP_LABELS.get(cur_group, ("KHÁC", "666666"))
        super_headers.append((start_col, len(out_keys), lbl, color))

    buf  = io.BytesIO()
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))

    with pd.ExcelWriter(buf, engine="openpyxl") as wr:
        df_ex.to_excel(wr, index=False, sheet_name="Tong_Hop", startrow=2)
        ws = wr.sheets["Tong_Hop"]

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(out_keys))
        tc = ws.cell(row=1, column=1)
        tc.value     = f"BẢNG TỔNG HỢP BÁO CÁO TGDV — KỲ: {ky_label.upper()}"
        tc.font      = Font(bold=True, size=13, color="FFFFFF")
        tc.fill      = PatternFill(start_color="1F3A5F", end_color="1F3A5F", fill_type="solid")
        tc.alignment = Alignment(horizontal="center", vertical="center")
        tc.border    = thin
        ws.row_dimensions[1].height = 28

        for start_c, end_c, lbl, color in super_headers:
            ws.merge_cells(start_row=2, start_column=start_c, end_row=2, end_column=end_c)
            cell = ws.cell(row=2, column=start_c)
            cell.value     = lbl
            cell.fill      = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            for c in range(start_c, end_c + 1):
                ws.cell(row=2, column=c).border = thin
        ws.row_dimensions[2].height = 22

        for ci in range(1, len(out_keys) + 1):
            cell = ws.cell(row=3, column=ci)
            cell.fill      = PatternFill(start_color="DCE4EF", end_color="DCE4EF", fill_type="solid")
            cell.font      = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = thin
        ws.row_dimensions[3].height = 38

        last_data_row = ws.max_row
        for ri in range(4, last_data_row + 1):
            is_total = ws.cell(row=ri, column=1).value == "TỔNG CỘNG"
            for ci in range(1, len(out_keys) + 1):
                cell = ws.cell(row=ri, column=ci)
                cell.border    = thin
                cell.alignment = Alignment(vertical="center", wrap_text=True,
                                           horizontal="center" if ci > 4 else "left")
                if is_total:
                    cell.font = Font(bold=True, color="A0443A", size=10)
                    cell.fill = PatternFill(start_color="F7EFDD", end_color="F7EFDD", fill_type="solid")
                elif ri % 2 == 0:
                    cell.fill = PatternFill(start_color="F7FAFD", end_color="F7FAFD", fill_type="solid")

        for ci, col_cells in enumerate(ws.columns, 1):
            lengths = []
            for cell in col_cells:
                try:
                    val = str(cell.value or "")
                    lengths.append(max(len(line) for line in val.split("\n")) if "\n" in val else len(val))
                except:
                    lengths.append(0)
            ws.column_dimensions[get_column_letter(ci)].width = min(max(max(lengths, default=6) + 2, 10), 42)

        ws.freeze_panes = "B4"
    return buf

# ==========================================
# KẾT NỐI SUPABASE (qua st.secrets)
# ==========================================
@st.cache_resource(show_spinner=False)
def get_supabase_client():
    try:
        from supabase import create_client
        url = st.secrets["supabase"]["url"]
        key = st.secrets["supabase"]["key"]
        return create_client(url, key)
    except Exception:
        return None


supabase = get_supabase_client()
SUPABASE_OK = supabase is not None


def log_access(app_name):
    key_name = f"da_dem_truy_cap_{app_name}"
    if SUPABASE_OK and key_name not in st.session_state:
        try:
            supabase.table("thong_ke_truy_cap").insert({"ten_app": app_name}).execute()
            st.session_state[key_name] = True
        except Exception:
            pass


log_access("Thu thập Báo cáo v2")


# ==========================================
# ĐĂNG NHẬP VÀ PHÂN QUYỀN
# ==========================================

def _get_admin_passwords():
    """Đọc mật khẩu quản trị / chuyên viên từ st.secrets.
    Cấu trúc secrets.toml mong đợi:
    [admin_passwords]
    "TongHop@2026" = "chuyen_vien"
    "Admin@2026" = "admin"
    """
    try:
        return dict(st.secrets["admin_passwords"])
    except Exception:
        # Dự phòng nếu chưa cấu hình secrets (cảnh báo nhẹ, không chặn ứng dụng)
        return {
            "TongHop@2026": "chuyen_vien",
            "Admin@2026":   "admin",
        }


PASSWORDS_QUAN_TRI = _get_admin_passwords()

ROLE_LABELS = {
    "user":        ("CƠ SỞ",                "🏘"),
    "chuyen_vien": ("CÁN BỘ TỔNG HỢP",      "📋"),
    "admin":       ("QUẢN TRỊ VIÊN",         "⚙"),
}

if "role" not in st.session_state:
    st.session_state.role = None
if "unit" not in st.session_state:
    st.session_state.unit = None

if st.session_state.role is None:
    col_l, col_m, col_r = st.columns([1, 1.1, 1])
    with col_m:
        st.markdown("""
        <div class='login-wrap'>
          <div class='login-logo'>
            <div style='font-size:46px;'>📊</div>
            <h2>HỆ THỐNG BÁO CÁO TGDV</h2>
            <p>Ban Tuyên giáo và Dân vận Tỉnh ủy — Vui lòng nhập mật khẩu để truy cập</p>
          </div>
        </div>
        """, unsafe_allow_html=True)
        with st.form("login_form"):
            pwd = st.text_input("Mật khẩu truy cập", type="password", placeholder="Nhập mật khẩu...")
            if st.form_submit_button("ĐĂNG NHẬP", use_container_width=True, type="primary"):
                # Check Admin / Cán bộ tổng hợp
                role = PASSWORDS_QUAN_TRI.get(pwd)
                if role:
                    st.session_state.role = role
                    st.session_state.unit = "ALL"
                    st.rerun()
                else:
                    # Check Pass Cơ sở
                    pwds = load_passwords()
                    matched_unit = None
                    for u, p in pwds.items():
                        if p == pwd:
                            matched_unit = u
                            break
                    if matched_unit:
                        st.session_state.role = "user"
                        st.session_state.unit = matched_unit
                        st.rerun()
                    else:
                        st.error("Mật khẩu không đúng. Vui lòng liên hệ Quản trị viên để được hỗ trợ.")
    st.stop()


# ==========================================
# ► KHỞI TẠO SESSION STATE BỘ LỌC
# ==========================================
if "filter_ky" not in st.session_state:
    st.session_state.filter_ky = "Tháng"
if "filter_thang" not in st.session_state:
    st.session_state.filter_thang = DANH_SACH_THANG[0]


# ==========================================
# SIDEBAR
# ==========================================
with st.sidebar:
    role_label, role_icon = ROLE_LABELS.get(st.session_state.role, ("?", "?"))

    hien_thi_ten = st.session_state.unit if st.session_state.role == "user" else role_label

    st.markdown(f"""
    <div style='background:rgba(255,255,255,0.08); border:1px solid rgba(255,255,255,0.18);
                border-radius:8px; padding:14px; text-align:center; margin-bottom:16px;'>
        <div style='font-size:26px;'>{role_icon}</div>
        <div style='font-size:11px; color:#A9BCD4; text-transform:uppercase; letter-spacing:1px; margin-top:4px;'>Đơn vị / Chức vụ</div>
        <div style='font-size:14px; font-weight:700; margin-top:2px;'>{hien_thi_ten}</div>
    </div>
    """, unsafe_allow_html=True)

    if st.session_state.role in ("admin", "chuyen_vien"):
        st.markdown("##### BỘ LỌC DỮ LIỆU")
        st.session_state.filter_ky = st.selectbox(
            "Kỳ tổng hợp",
            ["Tháng", "Quý I", "Quý II", "Quý III", "Quý IV",
             "6 Tháng Đầu Năm", "6 Tháng Cuối Năm", "9 Tháng", "Cả Năm"],
            index=["Tháng", "Quý I", "Quý II", "Quý III", "Quý IV",
                   "6 Tháng Đầu Năm", "6 Tháng Cuối Năm", "9 Tháng", "Cả Năm"
                   ].index(st.session_state.filter_ky),
            label_visibility="collapsed", key="sb_filter_ky"
        )

        if st.session_state.filter_ky == "Tháng":
            st.session_state.filter_thang = st.selectbox(
                "Chọn tháng", DANH_SACH_THANG,
                index=DANH_SACH_THANG.index(st.session_state.filter_thang),
                label_visibility="collapsed", key="sb_filter_thang"
            )
        st.markdown("---")

    # TÍNH NĂNG ĐỔI MẬT KHẨU CHO CƠ SỞ
    if st.session_state.role == "user":
        with st.expander("Đổi mật khẩu", expanded=False):
            with st.form("form_doi_pass", clear_on_submit=True):
                old_p = st.text_input("Mật khẩu hiện tại:", type="password")
                new_p = st.text_input("Mật khẩu mới:", type="password")
                cf_p  = st.text_input("Xác nhận mật khẩu mới:", type="password")

                if st.form_submit_button("Cập nhật", type="primary"):
                    pwds = load_passwords()
                    if pwds.get(st.session_state.unit) != old_p:
                        st.error("Mật khẩu hiện tại không đúng.")
                    elif new_p != cf_p:
                        st.error("Xác nhận mật khẩu không khớp.")
                    elif len(new_p) < 6:
                        st.error("Mật khẩu mới phải có ít nhất 6 ký tự.")
                    elif new_p in pwds.values() or new_p in PASSWORDS_QUAN_TRI.keys():
                        st.error("Mật khẩu này đã được sử dụng. Vui lòng chọn mật khẩu khác.")
                    else:
                        pwds[st.session_state.unit] = new_p
                        save_passwords(pwds)
                        st.success("Đổi mật khẩu thành công.")

    if st.button("Đăng xuất", use_container_width=True):
        st.session_state.role = None
        st.session_state.unit = None
        st.rerun()

    st.markdown(
        f"<div style='font-size:10px; color:#7691B0; text-align:center; margin-top:12px;'>"
        f"Phiên bản 5.0 · {datetime.now().strftime('%d/%m/%Y')}</div>",
        unsafe_allow_html=True
    )


# ==========================================
# HEADER CHÍNH
# ==========================================
st.markdown(
    "<div class='main-header'>HỆ THỐNG THU THẬP BÁO CÁO <span>CƠ SỞ TGDV</span></div>",
    unsafe_allow_html=True
)

# ==========================================
# TẠO TABS THEO QUYỀN
# ==========================================
if st.session_state.role == "admin":
    tab_nhap, tab_tiendo, tab_bieudo, tab_admin = st.tabs([
        "Nhập Báo cáo", "Tiến độ & Trích xuất", "Phân tích Biểu đồ", "Quản trị"
    ])
elif st.session_state.role == "chuyen_vien":
    tab_nhap, tab_tiendo, tab_bieudo = st.tabs([
        "Nhập Báo cáo", "Tiến độ & Trích xuất", "Phân tích Biểu đồ"
    ])
else:
    tab_nhap = st.tabs(["Nhập Báo cáo"])[0]


# ==========================================
# TAB 1: NHẬP BÁO CÁO
# ==========================================
with tab_nhap:
    st.markdown("<div class='section-title'>Xác định đơn vị và kỳ báo cáo</div>", unsafe_allow_html=True)
    c1, c2, c3 = st.columns([2.5, 1.5, 1.5])

    if st.session_state.role == "user":
        dv = st.session_state.unit
        c1.text_input("Đơn vị báo cáo", value=dv, disabled=True)
    else:
        dv = c1.selectbox("Đơn vị báo cáo", load_units(), index=None, placeholder="Gõ tên để tìm kiếm đơn vị...")

    nguoi_bc = c2.text_input("Người báo cáo / Số điện thoại", placeholder="Họ tên - Số điện thoại")
    th       = c3.selectbox("Kỳ báo cáo (tháng)", DANH_SACH_THANG, index=None, placeholder="Chọn tháng...")

    old = {}
    if dv and th:
        for d in load_data():
            if d.get("don_vi") == dv and d.get("ky_bao_cao") == th:
                old = d
                break

    if not dv or not th:
        st.markdown("""
        <div class='warning-box'>Vui lòng chọn <strong>Đơn vị</strong> và <strong>Kỳ báo cáo</strong> trước khi nhập số liệu.</div>
        """, unsafe_allow_html=True)
    else:
        if old:
            st.markdown(f"""<div class='info-box'>Đã tìm thấy báo cáo trước của <strong>{dv}</strong> — {th}. Số liệu đã được điền sẵn để chỉnh sửa.</div>""", unsafe_allow_html=True)
        else:
            st.markdown(f"""<div class='success-box'>Đang tạo báo cáo <strong>mới</strong> cho <strong>{dv}</strong> — {th}.</div>""", unsafe_allow_html=True)

        with st.form("f_nhap", clear_on_submit=False):

            with st.expander("1. Công tác lãnh đạo, chỉ đạo", expanded=True):
                c1, c2, c3 = st.columns(3)
                ld_vb = c1.number_input("Số VB cấp ủy ban hành",    min_value=0, value=get_old_val(old, "ld_vanban"))
                ld_tm = c2.number_input("Số VB tham mưu cấp trên",  min_value=0, value=get_old_val(old, "ld_thammuu"))
                ld_ch = c3.number_input("Số cuộc họp, hội nghị",    min_value=0, value=get_old_val(old, "ld_cuochop"))

            with st.expander("2. Học tập, quán triệt nghị quyết"):
                c1, c2, c3, c4 = st.columns(4)
                nq_hn = c1.number_input("Số hội nghị NQ",           min_value=0,   value=get_old_val(old, "nq_hoinghi"))
                nq_ng = c2.number_input("Số người tham gia",         min_value=0,   value=get_old_val(old, "nq_nguoi"))
                nq_vb = c3.number_input("Số VB đã triển khai",       min_value=0,   value=get_old_val(old, "nq_vanban"))
                nq_tl = c4.number_input("Tỷ lệ ĐV tham gia (%)",    min_value=0.0, max_value=100.0, step=0.5,
                                        value=get_old_val(old, "nq_tyle", 0.0))

            with st.expander("3. Công tác tuyên truyền"):
                c1, c2, c3, c4 = st.columns(4)
                tt_tb  = c1.number_input("Tin, bài, pano",           min_value=0, value=get_old_val(old, "tt_tinbai"))
                tt_lo  = c2.number_input("Lượt loa truyền thanh",    min_value=0, value=get_old_val(old, "tt_loa"))
                tt_bu  = c3.number_input("Buổi tuyên truyền miệng",            min_value=0, value=get_old_val(old, "tt_buoi"))
                tt_nn  = c4.number_input("Người nghe TT miệng",      min_value=0, value=get_old_val(old, "tt_nguoi"))
                c5, c6 = st.columns(2)
                tt_mxh  = c5.number_input("Bài trên MXH / Cổng TT", min_value=0, value=get_old_val(old, "tt_mxh_bai"))
                tt_ttmxh = c6.number_input("Lượt tương tác MXH",    min_value=0, value=get_old_val(old, "tt_mxh_tuongtac"))

            with st.expander("4. Nắm bắt dư luận xã hội"):
                c1, c2, c3 = st.columns(3)
                dl_bc = c1.number_input("Báo cáo DLXH gửi đi",      min_value=0, value=get_old_val(old, "dl_baocao"))
                dl_vd = c2.number_input("Vấn đề nổi cộm",            min_value=0, value=get_old_val(old, "dl_vande"))
                dl_xl = c3.number_input("Vụ việc đã xử lý",           min_value=0, value=get_old_val(old, "dl_xuly"))

            with st.expander("5. Khoa giáo, văn hóa - văn nghệ"):
                c1, c2, c3, c4 = st.columns(4)
                kg_ct = c1.number_input("CT tuyên truyền GD",        min_value=0, value=get_old_val(old, "kg_chuongtrinh"))
                kg_lo = c2.number_input("Buổi y tế / môi trường",      min_value=0, value=get_old_val(old, "kg_lop"))
                kg_cn = c3.number_input("Biểu diễn NT chuyên nghiệp",min_value=0, value=get_old_val(old, "kg_bd_chuyennghiep"))
                kg_qc = c4.number_input("Biểu diễn NT quần chúng",   min_value=0, value=get_old_val(old, "kg_bd_quanchung"))
                c5, c6, c7, c8 = st.columns(4)
                kg_cl = c5.number_input("CLB VH-NT thành lập",       min_value=0, value=get_old_val(old, "kg_clb_thanhlap"))
                kg_tv = c6.number_input("Thành viên CLB",            min_value=0, value=get_old_val(old, "kg_clb_thanhvien"))
                kg_lh = c7.number_input("HĐ lễ hội, thể thao",      min_value=0, value=get_old_val(old, "kg_hd_vhtt"))
                kg_hd = c8.number_input("HĐ văn hóa - văn nghệ",      min_value=0, value=get_old_val(old, "kg_hoatdong"))
                kg_kk = st.text_area("Khó khăn, vướng mắc (Khoa giáo, VH-VN):", value=get_old_val(old, "kg_khokhan"), height=80)

            with st.expander("6. Công tác dân vận (Dân vận khéo)"):
                c1, c2, c3 = st.columns(3)
                dv_dk  = c1.number_input("Mô hình DVK đăng ký",     min_value=0, value=get_old_val(old, "dv_mh_dangky"))
                dv_hq  = c2.number_input("Mô hình DVK hiệu quả",     min_value=0, value=get_old_val(old, "dv_mh_hieuqua"))
                dv_moi = c3.number_input("Mô hình mới trong kỳ",     min_value=0, value=get_old_val(old, "dv_mh_moi"))
                c4, c5, c6 = st.columns(3)
                dv_cv  = c4.number_input("Số cuộc vận động, TT",     min_value=0, value=get_old_val(old, "dv_cuocvandong"))
                dv_ntg = c5.number_input("Lượt người tham gia",      min_value=0, value=get_old_val(old, "dv_nguoithamgia"))
                dv_tx  = c6.number_input("Buổi đối thoại nhân dân",  min_value=0, value=get_old_val(old, "dv_tiepxuc"))

            with st.expander("7. Nhiệm vụ trọng tâm"):
                c1, c2, c3 = st.columns(3)
                nv_dg = c1.number_input("Nhiệm vụ TT được giao",     min_value=0, value=get_old_val(old, "nv_duocgiao"))
                nv_ht = c2.number_input("Nhiệm vụ TT hoàn thành",    min_value=0, value=get_old_val(old, "nv_hoanthanh"))
                nv_dk = c3.number_input("Nhiệm vụ đang triển khai",  min_value=0, value=get_old_val(old, "nv_dangtrienkhai"))
                nv_kq = st.text_area("Kết quả thí điểm nổi bật:", value=get_old_val(old, "nv_ketqua"), height=80)

            with st.expander("8. Chuyên đề: Bình dân học vụ số"):
                st.markdown("**Thông tin chung**")
                c1, c2, c3 = st.columns(3)
                bd_ti = c1.number_input("Tin bài về CĐS",            min_value=0, value=get_old_val(old, "bd_tinbai"))
                bd_ct = c2.number_input("Cuộc thi CĐS",              min_value=0, value=get_old_val(old, "bd_cuocthi"))
                kq_tc = c3.number_input("Số Tổ công nghệ số",        min_value=0, value=get_old_val(old, "kq_tocongnghe"))

                st.markdown("**Đối với chi bộ**")
                c4, c5, c6 = st.columns(3)
                ts_ch = c4.number_input("Tổng số chi bộ",            min_value=0, value=get_old_val(old, "ts_chibo"))
                kq_cd = c5.number_input("CB sinh hoạt chuyên đề số",        min_value=0, value=get_old_val(old, "kq_chibo_cd"))
                kq_st = c6.number_input("CB dùng sổ tay ĐV số",      min_value=0, value=get_old_val(old, "kq_chibo_sotay"))

                st.markdown("**Đối với cán bộ, công chức, viên chức**")
                c7, c8, c9 = st.columns(3)
                ts_cb = c7.number_input("Tổng số CBCCVC",            min_value=0, value=get_old_val(old, "ts_cbccvc"))
                kq_ai = c8.number_input("CB biết dùng AI",           min_value=0, value=get_old_val(old, "kq_cb_ai"))
                kq_ck = c9.number_input("CB hoàn thành khóa CĐS",   min_value=0, value=get_old_val(old, "kq_cb_khoahoc"))

                st.markdown("**Đối với nhân dân**")
                c10, c11, c12 = st.columns(3)
                ts_nd = c10.number_input("Tổng người dân trưởng thành",     min_value=0, value=get_old_val(old, "ts_nd_truongthanh"))
                kq_kn = c11.number_input("Người dân có kỹ năng số",         min_value=0, value=get_old_val(old, "kq_nd_kynang"))
                kq_vi = c12.number_input("Người dân phổ cập VNeID",         min_value=0, value=get_old_val(old, "kq_nd_vneid"))
                c13, c14 = st.columns(2)
                kq_sm = c13.number_input("Người dân dùng smartphone",       min_value=0, value=get_old_val(old, "kq_nd_smartphone"))
                kq_lp = c14.number_input("Buổi học cộng đồng",       min_value=0, value=get_old_val(old, "kq_lop_nd"))

            with st.expander("9. Đánh giá chung & kiến nghị"):
                tl_mo = st.text_area("Mô hình hay, sáng tạo trong tháng:", value=get_old_val(old, "tl_mohinh"), height=100,
                                     placeholder="Mô tả các mô hình, cách làm hay, sáng tạo...")
                tl_kh = st.text_area("Khó khăn, vướng mắc và kiến nghị:", value=get_old_val(old, "tl_khokhan"), height=100,
                                     placeholder="Nêu rõ khó khăn và đề xuất hướng giải quyết...")

            st.markdown("")
            submitted = st.form_submit_button("GỬI / CẬP NHẬT BÁO CÁO", use_container_width=True, type="primary")

            if submitted:
                if not nguoi_bc.strip():
                    st.error("Vui lòng điền tên người báo cáo và số điện thoại.")
                else:
                    new_rec = {
                        "don_vi": dv, "ky_bao_cao": th,
                        "nguoi_bao_cao": nguoi_bc.strip(),
                        "ngay_nop": datetime.now().strftime("%d/%m/%Y %H:%M"),
                        "ld_vanban": ld_vb, "ld_thammuu": ld_tm, "ld_cuochop": ld_ch,
                        "nq_hoinghi": nq_hn, "nq_nguoi": nq_ng, "nq_vanban": nq_vb, "nq_tyle": nq_tl,
                        "tt_tinbai": tt_tb, "tt_loa": tt_lo, "tt_buoi": tt_bu, "tt_nguoi": tt_nn,
                        "tt_mxh_bai": tt_mxh, "tt_mxh_tuongtac": tt_ttmxh,
                        "dl_baocao": dl_bc, "dl_vande": dl_vd, "dl_xuly": dl_xl,
                        "kg_chuongtrinh": kg_ct, "kg_lop": kg_lo, "kg_bd_chuyennghiep": kg_cn,
                        "kg_bd_quanchung": kg_qc, "kg_clb_thanhlap": kg_cl, "kg_clb_thanhvien": kg_tv,
                        "kg_hd_vhtt": kg_lh, "kg_hoatdong": kg_hd, "kg_khokhan": kg_kk,
                        "dv_mh_dangky": dv_dk, "dv_mh_hieuqua": dv_hq, "dv_mh_moi": dv_moi,
                        "dv_cuocvandong": dv_cv, "dv_nguoithamgia": dv_ntg, "dv_tiepxuc": dv_tx,
                        "nv_duocgiao": nv_dg, "nv_hoanthanh": nv_ht, "nv_dangtrienkhai": nv_dk, "nv_ketqua": nv_kq,
                        "bd_tinbai": bd_ti, "bd_cuocthi": bd_ct, "kq_tocongnghe": kq_tc,
                        "ts_chibo": ts_ch, "kq_chibo_cd": kq_cd, "kq_chibo_sotay": kq_st,
                        "ts_cbccvc": ts_cb, "kq_cb_ai": kq_ai, "kq_cb_khoahoc": kq_ck,
                        "ts_nd_truongthanh": ts_nd, "kq_nd_kynang": kq_kn, "kq_nd_vneid": kq_vi,
                        "kq_nd_smartphone": kq_sm, "kq_lop_nd": kq_lp,
                        "tl_mohinh": tl_mo, "tl_khokhan": tl_kh,
                    }
                    data = load_data()
                    data = [d for d in data if not (d.get("don_vi") == dv and d.get("ky_bao_cao") == th)]
                    data.append(new_rec)
                    save_data(data)
                    st.success(f"Đã lưu thành công báo cáo của {dv} — {th}.")
                    st.balloons()
                    old = new_rec

        if st.session_state.role == "user" and old:
            st.markdown("<div style='height: 15px'></div>", unsafe_allow_html=True)
            df_unit = pd.DataFrame([old])
            df_unit_sum = build_summary_df(df_unit, th)
            if not df_unit_sum.empty:
                excel_bytes_unit = build_excel_cached(
                    df_unit_sum.to_json(orient="split"), th
                )
                st.download_button(
                    label=f"Tải báo cáo ({th}) để lưu trữ",
                    data=excel_bytes_unit,
                    file_name=f"Bao_cao_{dv}_{th.replace(' ', '_')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="secondary",
                    use_container_width=True
                )

# ==========================================
# TAB 2 & 3: CHỈ HIỂN THỊ VỚI ADMIN / CHUYÊN VIÊN
# ==========================================
if st.session_state.role in ("admin", "chuyen_vien"):

    l_bc = st.session_state.filter_ky
    if l_bc == "Tháng":
        months_sel = [st.session_state.filter_thang]
        ky_label   = st.session_state.filter_thang
    else:
        months_sel = get_months_for_filter(l_bc)
        ky_label   = l_bc

    all_raw = load_data()
    all_u   = load_units()

    df_cur  = pd.DataFrame([d for d in all_raw if d.get("ky_bao_cao") in months_sel])
    u_done  = df_cur["don_vi"].unique().tolist() if not df_cur.empty else []
    u_miss  = [u for u in all_u if u not in u_done]
    rate    = (len(u_done) / len(all_u) * 100) if all_u else 0

    if not df_cur.empty:
        df_sum = build_summary_df_cached(df_cur.to_json(orient="records"), ky_label)
    else:
        df_sum = pd.DataFrame()

    # ── TAB 2: TIẾN ĐỘ & TRÍCH XUẤT ──
    with tab_tiendo:
        st.markdown(f"<div class='section-title'>Tiến độ nộp báo cáo — Kỳ: {ky_label.upper()}</div>",
                    unsafe_allow_html=True)

        c1, c2, c3, c4 = st.columns(4)
        c1.markdown(f"""<div class='metric-card green'><div class='metric-title'>Đã nộp</div><div class='metric-number green'>{len(u_done)}</div><div class='metric-sub'>đơn vị</div></div>""", unsafe_allow_html=True)
        c2.markdown(f"""<div class='metric-card red'><div class='metric-title'>Chưa nộp</div><div class='metric-number red'>{len(u_miss)}</div><div class='metric-sub'>đơn vị</div></div>""", unsafe_allow_html=True)
        c3.markdown(f"""<div class='metric-card' style='border-left-color:#2C4A75;'><div class='metric-title'>Tỷ lệ hoàn thành</div><div class='metric-number' style='color:#2C4A75;'>{rate:.1f}%</div><div class='metric-sub'>tiến độ</div></div>""", unsafe_allow_html=True)
        c4.markdown(f"""<div class='metric-card gold'><div class='metric-title'>Tổng đơn vị</div><div class='metric-number gold'>{len(all_u)}</div><div class='metric-sub'>trong hệ thống</div></div>""", unsafe_allow_html=True)

        bar_color = "#3E7D5A" if rate >= 80 else ("#B08D57" if rate >= 50 else "#A0443A")
        st.markdown(f"""
        <div style='background:white; padding:16px 20px; border-radius:8px; margin:16px 0; box-shadow:0 1px 6px rgba(0,0,0,0.04);'>
            <div style='display:flex; justify-content:space-between; font-size:13px; font-weight:600; color:#3C4856; margin-bottom:8px;'>
                <span>Tiến độ nộp báo cáo</span>
                <span style='color:{bar_color};'>{rate:.1f}%</span>
            </div>
            <div class='progress-wrap'>
                <div class='progress-fill' style='width:{rate:.1f}%;'></div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        if not df_sum.empty:
            excel_bytes = build_excel_cached(df_sum.to_json(orient="split"), ky_label)
            st.download_button(
                label=f"Tải bảng tổng hợp số liệu — {ky_label.upper()} (Excel)",
                data=excel_bytes,
                file_name=f"TongHop_BaoCao_{ky_label.replace(' ', '_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary", use_container_width=True
            )
        else:
            st.info(f"Chưa có dữ liệu cho kỳ {ky_label}. Vui lòng kiểm tra lại bộ lọc.")

        st.markdown("<div class='section-title'>Chi tiết tiến độ theo đơn vị</div>", unsafe_allow_html=True)
        col_a, col_b = st.columns([1.2, 1])

        with col_a:
            st.markdown(f"<span class='badge badge-success'>Đã nộp — {len(u_done)} đơn vị</span>", unsafe_allow_html=True)
            if u_done and not df_cur.empty:
                df_d = (df_cur[["don_vi", "ngay_nop", "nguoi_bao_cao"]]
                        .drop_duplicates(subset=["don_vi"], keep="last")
                        .sort_values("don_vi")
                        .rename(columns={"don_vi": "Đơn vị", "ngay_nop": "Thời điểm nộp", "nguoi_bao_cao": "Người nộp"}))
                st.dataframe(df_d, use_container_width=True, hide_index=True, height=400)
            else:
                st.info("Chưa có đơn vị nào nộp báo cáo.")

        with col_b:
            st.markdown(f"<span class='badge badge-danger'>Chưa nộp — {len(u_miss)} đơn vị</span>", unsafe_allow_html=True)
            if u_miss:
                for m in sorted(u_miss):
                    st.markdown(f"<div style='padding:5px 0; border-bottom:1px solid #F0F4F8; font-size:13px;'>{m}</div>", unsafe_allow_html=True)
            else:
                st.success("Tất cả đơn vị đã nộp báo cáo.")

    # ── TAB 3: BIỂU ĐỒ ──
    with tab_bieudo:
        if df_sum.empty:
            st.warning("Chưa có số liệu để vẽ biểu đồ. Vui lòng kiểm tra lại bộ lọc.")
        else:
            df_plot = df_sum[df_sum["don_vi"] != "TỔNG CỘNG"].copy()
            st.info(f"Đang hiển thị phân tích số liệu kỳ {ky_label} — {len(u_done)} đơn vị đã nộp")

            COLORS = ["#1F3A5F", "#A0443A", "#3E7D5A", "#B08D57", "#6B5B95", "#3F8FA8"]

            def make_bar(df, x, ys, names, title, colors=None):
                fig = go.Figure()
                for i, (y, n) in enumerate(zip(ys, names)):
                    if y in df.columns:
                        fig.add_trace(go.Bar(
                            name=n, x=df[x], y=df[y],
                            marker_color=(colors or COLORS)[i % len(COLORS)],
                            text=df[y], textposition="outside", textfont_size=9
                        ))
                fig.update_layout(
                    title=dict(text=title, font=dict(size=13, color="#1F3A5F")),
                    barmode="group", height=360,
                    plot_bgcolor="white", paper_bgcolor="white",
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1, font_size=11),
                    xaxis=dict(tickangle=-30, tickfont_size=9),
                    margin=dict(t=60, b=60, l=40, r=20)
                )
                return fig

            st.markdown("<div class='section-title'>1. Lãnh đạo & 2. Nghị quyết</div>", unsafe_allow_html=True)
            r1c1, r1c2 = st.columns(2)
            with r1c1:
                st.plotly_chart(make_bar(df_plot, "don_vi",
                    ["ld_vanban", "ld_thammuu", "ld_cuochop"],
                    ["VB ban hành", "VB tham mưu", "Cuộc họp"],
                    "Công tác lãnh đạo, chỉ đạo"), use_container_width=True)
            with r1c2:
                st.plotly_chart(make_bar(df_plot, "don_vi",
                    ["nq_hoinghi", "nq_vanban", "nq_nguoi"],
                    ["Hội nghị NQ", "VB triển khai", "Người tham gia"],
                    "Quán triệt nghị quyết"), use_container_width=True)

            st.markdown("<div class='section-title'>3. Tuyên truyền & 4. Dư luận xã hội</div>", unsafe_allow_html=True)
            r2c1, r2c2 = st.columns(2)
            with r2c1:
                st.plotly_chart(make_bar(df_plot, "don_vi",
                    ["tt_tinbai", "tt_mxh_bai", "tt_buoi"],
                    ["Tin bài", "Bài MXH", "Buổi TT miệng"],
                    "Công tác tuyên truyền"), use_container_width=True)
            with r2c2:
                dl_vals = [
                    df_plot["dl_baocao"].sum() if "dl_baocao" in df_plot.columns else 0,
                    df_plot["dl_vande"].sum()  if "dl_vande"  in df_plot.columns else 0,
                    df_plot["dl_xuly"].sum()   if "dl_xuly"   in df_plot.columns else 0,
                ]
                fig_pie = go.Figure(go.Pie(
                    labels=["Báo cáo gửi đi", "Vấn đề nổi cộm", "Đã xử lý"],
                    values=dl_vals, hole=0.5,
                    marker_colors=["#1F3A5F", "#A0443A", "#3E7D5A"], textfont_size=11
                ))
                fig_pie.update_layout(
                    title=dict(text="Tổng hợp dư luận xã hội", font=dict(size=13, color="#1F3A5F")),
                    height=360, paper_bgcolor="white", legend=dict(font_size=11), margin=dict(t=60, b=20)
                )
                st.plotly_chart(fig_pie, use_container_width=True)

            st.markdown("<div class='section-title'>5. Khoa giáo & 6. Dân vận khéo</div>", unsafe_allow_html=True)
            r3c1, r3c2 = st.columns(2)
            with r3c1:
                st.plotly_chart(make_bar(df_plot, "don_vi",
                    ["kg_chuongtrinh", "kg_bd_chuyennghiep", "kg_clb_thanhlap"],
                    ["CT tuyên truyền GD", "Biểu diễn NT", "CLB thành lập"],
                    "Khoa giáo, văn hóa - văn nghệ"), use_container_width=True)
            with r3c2:
                st.plotly_chart(make_bar(df_plot, "don_vi",
                    ["dv_mh_dangky", "dv_mh_hieuqua", "dv_mh_moi"],
                    ["MH đăng ký", "MH hiệu quả", "MH mới"],
                    "Phong trào dân vận khéo"), use_container_width=True)

            st.markdown("<div class='section-title'>7. Nhiệm vụ trọng tâm & 8. Bình dân học vụ số</div>", unsafe_allow_html=True)
            r4c1, r4c2 = st.columns(2)
            with r4c1:
                st.plotly_chart(make_bar(df_plot, "don_vi",
                    ["nv_duocgiao", "nv_hoanthanh", "nv_dangtrienkhai"],
                    ["Được giao", "Hoàn thành", "Đang triển khai"],
                    "Tiến độ nhiệm vụ trọng tâm"), use_container_width=True)
            with r4c2:
                st.plotly_chart(make_bar(df_plot, "don_vi",
                    ["kq_cb_ai", "kq_nd_kynang", "kq_nd_vneid"],
                    ["CB dùng AI", "ND kỹ năng số", "ND VNeID"],
                    "Bình dân học vụ số — Kết quả nổi bật"), use_container_width=True)

            st.markdown("<div class='section-title'>Tỷ lệ đảng viên tham gia nghị quyết (%)</div>", unsafe_allow_html=True)
            if "nq_tyle" in df_plot.columns:
                fig_rate = px.bar(
                    df_plot.sort_values("nq_tyle", ascending=True),
                    x="nq_tyle", y="don_vi", orientation="h",
                    color="nq_tyle", color_continuous_scale=["#A0443A", "#B08D57", "#3E7D5A"],
                    range_color=[0, 100],
                    title="Tỷ lệ đảng viên tham gia học nghị quyết (%)",
                    labels={"nq_tyle": "Tỷ lệ (%)", "don_vi": ""},
                    text="nq_tyle"
                )
                fig_rate.update_traces(texttemplate="%{text:.1f}%", textposition="outside")
                fig_rate.update_layout(
                    height=max(300, len(df_plot) * 28),
                    plot_bgcolor="white", paper_bgcolor="white",
                    yaxis=dict(tickfont_size=10),
                    margin=dict(l=180, r=60, t=60, b=20),
                    coloraxis_showscale=False
                )
                st.plotly_chart(fig_rate, use_container_width=True)

# ==========================================
# TAB 4: QUẢN TRỊ ADMIN
# ==========================================
if st.session_state.role == "admin":
    with tab_admin:
        st.markdown("<div class='section-title'>Quản trị hệ thống</div>", unsafe_allow_html=True)
        col_adm1, col_adm2 = st.columns(2)

        with col_adm1:
            st.markdown("#### Xóa báo cáo")
            cur_data = load_data()
            if cur_data:
                dv_list = sorted(set(d.get("don_vi") for d in cur_data if d.get("don_vi")))
                d_v = st.selectbox("Chọn đơn vị cần xóa:", dv_list, key="dv_del")

                if d_v:
                    th_list = sorted(set(d.get("ky_bao_cao") for d in cur_data if d.get("don_vi") == d_v and d.get("ky_bao_cao")))
                    if th_list:
                        t_h = st.selectbox("Chọn tháng:", th_list, key="th_del")
                        st.warning(f"Sẽ xóa báo cáo của {d_v} — {t_h}. Thao tác không thể hoàn tác.")
                        if st.button("XÁC NHẬN XÓA BÁO CÁO", type="primary"):
                            new_d = [d for d in cur_data if not (d.get("don_vi") == d_v and d.get("ky_bao_cao") == t_h)]
                            save_data(new_d)
                            st.success("Đã xóa báo cáo.")
                            st.rerun()
            else:
                st.info("Chưa có dữ liệu báo cáo nào.")

            st.markdown("---")
            st.markdown("#### Xuất toàn bộ dữ liệu thô (JSON)")
            if cur_data:
                st.download_button(
                    "Tải file JSON backup",
                    data=json.dumps(cur_data, ensure_ascii=False, indent=2),
                    file_name=f"backup_baocao_{datetime.now().strftime('%Y%m%d_%H%M')}.json",
                    mime="application/json"
                )

        with col_adm2:
            st.markdown("#### Quản lý mật khẩu cơ sở")
            st.info("Các đơn vị có thể tự đổi mật khẩu. Nếu đơn vị báo quên, Quản trị viên có thể khôi phục tại đây.")

            pwds = load_passwords()
            df_acc = pd.DataFrame([{"Đơn vị": k, "Mật khẩu": v} for k, v in pwds.items()])

            # --- Ô tìm kiếm mật khẩu ---
            search_pwd = st.text_input("Tìm kiếm mật khẩu theo tên đơn vị:", placeholder="Gõ tên xã/phường...")
            df_acc_display = df_acc.copy()
            if search_pwd:
                df_acc_display = df_acc_display[df_acc_display["Đơn vị"].str.contains(search_pwd, case=False, na=False)]

            st.dataframe(df_acc_display, hide_index=True, height=200)

            # --- Nút tải danh sách tài khoản CSV ---
            csv = df_acc.to_csv(index=False).encode('utf-8-sig')
            st.download_button(
                label="Tải danh sách tài khoản (CSV)",
                data=csv,
                file_name=f"Danh_sach_tai_khoan_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv",
                use_container_width=True
            )

            with st.form("form_reset_pass"):
                u_reset = st.selectbox("Chọn đơn vị cần khôi phục mật khẩu:", df_acc["Đơn vị"].tolist())
                if st.form_submit_button("Khôi phục mật khẩu mặc định", type="secondary"):
                    units = load_units()
                    idx = units.index(u_reset)
                    default_p = f"TGDV@{idx+1:03d}"
                    pwds[u_reset] = default_p
                    save_passwords(pwds)
                    st.success(f"Đã khôi phục mật khẩu của {u_reset} thành: {default_p}")
                    st.rerun()

            st.markdown("---")
            st.markdown("#### Quản lý danh sách đơn vị")
            u_list = load_units()

            with st.form("form_add_unit"):
                new_u = st.text_input("Tên đơn vị mới:")
                if st.form_submit_button("Thêm đơn vị", type="primary"):
                    if new_u.strip() and new_u.strip() not in u_list:
                        u_list.append(new_u.strip())
                        save_units(u_list)
                        st.success(f"Đã thêm: {new_u.strip()}")
                        st.rerun()
                    elif new_u.strip() in u_list:
                        st.warning("Đơn vị đã tồn tại trong danh sách.")
                    else:
                        st.error("Tên đơn vị không được để trống.")

            st.markdown("---")
            with st.form("form_del_unit"):
                rem_u = st.selectbox("Chọn đơn vị cần xóa:", ["-- Chọn đơn vị --"] + sorted(u_list))
                if st.form_submit_button("Xóa đơn vị", type="secondary"):
                    if rem_u != "-- Chọn đơn vị --":
                        u_list.remove(rem_u)
                        save_units(u_list)
                        st.success(f"Đã xóa: {rem_u}")
                        st.rerun()
                    else:
                        st.warning("Vui lòng chọn đơn vị cần xóa.")
